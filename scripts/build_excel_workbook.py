import pandas as pd
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = Path("data")
OUTPUT_DIR = Path("output")

HEADER_FILL = "1F4E78"
HEADER_FONT = "FFFFFF"

ACTION_FILLS = {
    "SELL": "F4CCCC",
    "TRIM": "FCE4D6",
    "BUY": "C6EFCE",
    "BUY SMALL": "D9EAF7",
    "HOLD": "E2F0D9",
    "WATCH": "FFF2CC",
    "TIMING CONFIRMED": "C6EFCE",
    "TIMING WATCH": "FFF2CC",
    "WAIT": "E7E6E6",
    "PROMOTE": "C6EFCE",
    "KEEP CANDIDATE": "FFF2CC",
    "REMOVE": "F4CCCC",
    "ALREADY ACTIVE": "D9EAF7",
}

QUALITY_FILLS = {
    "HIGH": "F4CCCC",
    "MEDIUM": "FFF2CC",
    "LOW": "D9EAF7",
    "OK": "C6EFCE",
}

CURRENCY_COLUMNS = {
    "buy_usd", "sell_usd", "price", "stop_loss", "trim_target",
    "avg_cost", "cost_basis", "current_price", "market_value",
    "unrealized_pnl", "total_deposits_nzd", "total_deposits_usd",
    "cash_available_usd", "realized_pnl_usd", "avg_pnl_usd",
    "buy_cost_usd", "sell_proceeds_usd", "sell_fee_usd",
    "buy_fees_usd", "value", "usd_amount", "transaction_fee",
}

PERCENT_COLUMNS = {
    "holding_return_pct", "win_rate", "avg_return_pct",
    "realized_return_pct", "gap_pct", "atr_pct",
}

INTEGER_COLUMNS = {
    "score", "closed_trades", "wins", "losses",
    "timing_score", "holding_days", "avg_holding_days",
    "trend_score", "momentum_score", "accelerator_score",
}

LONG_TEXT_COLUMNS = {
    "plan_why", "risk_note", "position_rule", "notes",
    "latest_reasons", "latest_warnings", "issue", "suggested_action",
    "position_rationale", "gap_reason", "volume_reason",
    "macd_reason", "rsi_reason", "data_note", "details", "Action",
    "why", "candidate_reason",
}

DISPLAY_OVERRIDES = {
    "usd": "USD",
    "nzd": "NZD",
    "pnl": "P&L",
    "rsi": "RSI",
    "macd": "MACD",
    "atr": "ATR",
    "sec": "SEC",
}

def read_csv_safe(path):
    try:
        if path.exists() and path.stat().st_size > 0:
            return pd.read_csv(path)
    except pd.errors.EmptyDataError:
        pass
    return pd.DataFrame()

def safe_sheet_name(name):
    return name[:31]

def prettify_header(header):
    parts = str(header).replace("_", " ").split()
    pretty = []

    for part in parts:
        key = part.lower()
        if key in DISPLAY_OVERRIDES:
            pretty.append(DISPLAY_OVERRIDES[key])
        else:
            pretty.append(part.capitalize())

    return " ".join(pretty)

def prettify_headers(ws):
    for cell in ws[1]:
        cell.value = prettify_header(cell.value)

def apply_table_style(ws):
    thin = Side(style="thin", color="D9EAD3")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
            cell.border = Border(bottom=thin)

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(color=HEADER_FONT, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

def apply_number_formats(ws, raw_headers):
    for col_idx, header in enumerate(raw_headers, start=1):
        header_key = str(header).strip()

        if header_key in CURRENCY_COLUMNS:
            fmt = "$#,##0.00"
        elif header_key in PERCENT_COLUMNS:
            fmt = "0.0%"
        elif header_key in INTEGER_COLUMNS:
            fmt = "0"
        elif header_key in {"shares", "shares_to_buy", "shares_to_sell"}:
            fmt = "0.000000"
        else:
            fmt = None

        if fmt:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = fmt

def apply_action_highlights(ws, raw_headers, action_header):
    if action_header not in raw_headers:
        return

    action_col = raw_headers.index(action_header) + 1

    for row in ws.iter_rows(min_row=2):
        action = str(row[action_col - 1].value).upper()
        fill_color = ACTION_FILLS.get(action)

        if fill_color:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=fill_color)

def apply_quality_highlights(ws, raw_headers):
    if "severity" not in raw_headers:
        return

    severity_col = raw_headers.index("severity") + 1

    for row in ws.iter_rows(min_row=2):
        severity = str(row[severity_col - 1].value).upper()
        fill_color = QUALITY_FILLS.get(severity)

        if fill_color:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=fill_color)

def set_column_widths(ws, raw_headers):
    for col_idx, header in enumerate(raw_headers, start=1):
        header_key = str(header).strip()
        col_letter = get_column_letter(col_idx)

        if header_key in LONG_TEXT_COLUMNS:
            ws.column_dimensions[col_letter].width = 42
            for row in range(1, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).alignment = Alignment(
                    horizontal="left",
                    vertical="top",
                    wrap_text=True
                )
        elif header_key in {
            "ticker", "tier", "action_required", "add_more",
            "timing_action", "severity", "recommendation"
        }:
            ws.column_dimensions[col_letter].width = 18
        elif header_key in CURRENCY_COLUMNS or header_key in PERCENT_COLUMNS:
            ws.column_dimensions[col_letter].width = 14
        elif header_key in {
            "date", "week_start", "sell_date", "first_buy_date",
            "as_of", "fetched_at_utc", "market_data_date", "date_added"
        }:
            ws.column_dimensions[col_letter].width = 18
        else:
            max_len = len(prettify_header(header_key))
            for row in range(2, min(ws.max_row, 60) + 1):
                value = ws.cell(row=row, column=col_idx).value
                max_len = max(max_len, len("" if value is None else str(value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 24)

    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_dashboard(ws, raw_headers):
    set_column_widths(ws, raw_headers)

    preferred = {
        "ticker": 12,
        "action_required": 14,
        "position_rule": 28,
        "plan_why": 48,
        "buy_usd": 12,
        "sell_usd": 12,
        "shares_to_buy": 14,
        "shares_to_sell": 14,
        "price": 12,
        "score": 10,
        "tier": 14,
        "holding_return_pct": 14,
        "stop_loss": 12,
        "trim_target": 12,
        "risk_note": 44,
        "add_more": 12,
    }

    for idx, header in enumerate(raw_headers, start=1):
        if header in preferred:
            ws.column_dimensions[get_column_letter(idx)].width = preferred[header]

    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def build_excel_workbook():
    OUTPUT_DIR.mkdir(exist_ok=True)

    files = {
        "Dashboard": DATA_DIR / "dashboard.csv",
        "Candidate Review": DATA_DIR / "candidate_review.csv",
        "Holdings": DATA_DIR / "holdings_view.csv",
        "Trades": DATA_DIR / "trades.csv",
        "Performance Summary": DATA_DIR / "performance_summary.csv",
        "Weekly Review": DATA_DIR / "weekly_review.csv",
        "Closed Trades Learning": DATA_DIR / "closed_trades_learning.csv",
        "Performance Learning": DATA_DIR / "performance_learning.csv",
        "Data Quality": DATA_DIR / "data_quality_report.csv",
        "Signals Full": DATA_DIR / "signals.csv",
        "Market Timing": DATA_DIR / "market_timing.csv",
        "Market Data": DATA_DIR / "market_data.csv",
    }

    excel_file = OUTPUT_DIR / "trading_summary.xlsx"
    raw_headers_by_sheet = {}

    with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
        for sheet_name, path in files.items():
            df = read_csv_safe(path)

            if df.empty:
                df = pd.DataFrame([{"status": "No data available yet"}])

            safe_name = safe_sheet_name(sheet_name)
            raw_headers_by_sheet[safe_name] = list(df.columns)
            df.to_excel(writer, sheet_name=safe_name, index=False)

        instructions = pd.DataFrame([
            {"Step": 1, "Action": "Open Dashboard first. Review held positions and the active Dashboard watchlist."},
            {"Step": 2, "Action": "Review SELL / TRIM rows before any buys."},
            {"Step": 3, "Action": "Check position_rule and plan_why together before trading."},
            {"Step": 4, "Action": "Dashboard watchlist source: data/dashboard_watchlist.csv."},
            {"Step": 5, "Action": "Candidate pool source: data/candidate_pool.csv. Candidate Review shows names not currently active on Dashboard."},
            {"Step": 6, "Action": "Check Holdings and Performance Summary before trading."},
            {"Step": 7, "Action": "Use BUY / BUY SMALL as candidates, not automatic trades."},
            {"Step": 8, "Action": "Check Market Timing to confirm whether entry timing looks supportive."},
            {"Step": 9, "Action": "Check Weekly Review, Closed Trades Learning, and Performance Learning."},
            {"Step": 10, "Action": "Check Data Quality if a signal looks strange."},
            {"Step": 11, "Action": "For full sells in trades.csv, use shares=ALL."},
            {"Step": 12, "Action": "Run GitHub workflow, then open the generated workbook."},
        ])

        raw_headers_by_sheet["How To Use"] = list(instructions.columns)
        instructions.to_excel(writer, sheet_name="How To Use", index=False)

        wb = writer.book

        preferred_order = [
            "Dashboard",
            "Candidate Review",
            "Holdings",
            "Performance Summary",
            "Data Quality",
            "Market Timing",
            "Weekly Review",
            "Closed Trades Learning",
            "Performance Learning",
            "Trades",
            "Signals Full",
            "Market Data",
            "How To Use",
        ]

        wb._sheets.sort(
            key=lambda ws: preferred_order.index(ws.title) if ws.title in preferred_order else 999
        )

        for ws in wb.worksheets:
            raw_headers = raw_headers_by_sheet.get(ws.title, [cell.value for cell in ws[1]])

            apply_table_style(ws)
            apply_number_formats(ws, raw_headers)
            set_column_widths(ws, raw_headers)

            if ws.title == "Dashboard":
                apply_action_highlights(ws, raw_headers, "action_required")
                style_dashboard(ws, raw_headers)

            if ws.title == "Candidate Review":
                apply_action_highlights(ws, raw_headers, "recommendation")

            if ws.title == "Signals Full":
                apply_action_highlights(ws, raw_headers, "action")

            if ws.title == "Market Timing":
                apply_action_highlights(ws, raw_headers, "timing_action")

            if ws.title == "Data Quality":
                apply_quality_highlights(ws, raw_headers)

            prettify_headers(ws)

            for row in range(1, ws.max_row + 1):
                ws.row_dimensions[row].height = 24 if row == 1 else None

            for cell in ws[1]:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    print(f"Excel workbook created: {excel_file}")
    return excel_file

if __name__ == "__main__":
    build_excel_workbook()
